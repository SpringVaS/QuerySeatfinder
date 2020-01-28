from printers import Printer

from openpyxl import Workbook, load_workbook

import pandas as pd

import myutils

class ExcelPrinter(Printer):

	def __init__(self, dstpath):
		self.dstpath = dstpath
		workbook = Workbook()
		workbook.save(dstpath)


	def export_lib_metadata(self, metadata):
		self.__write_to_excel(self.__displayable_lib_metadata(metadata), "Libraries", True)

	def export_data(self, data, title, quantity_description):
		self.__write_to_excel(data, name)

	def finish_up(self):
		self.__delete_standard_sheet()

	def __displayable_lib_metadata(self, metadata):
		displayable_metadata = metadata.copy(deep = True)
		for location_id in metadata.keys():
			entry = displayable_metadata[location_id]['opening_hours']
			ohlist_display = []
			for interval in entry:
				openingHours_str = ""
				weekday_opening = interval.left.strftime("%A")
				weekday_closing = interval.right.strftime("%A")
				time_opening = interval.left.strftime("%H:%M")
				time_closing = interval.right.strftime("%H:%M")
				if (weekday_opening == weekday_closing):
					openingHours_str = weekday_opening + ": " + time_opening + " - " + time_closing
				else:
					openingHours_str = weekday_opening + ": " + time_opening + " - " + weekday_closing + ": " + time_closing
				ohlist_display.append(openingHours_str)
			displayable_metadata[location_id]['opening_hours'] = ohlist_display

		displayable_metadata = pd.merge(displayable_metadata[self.mainlib].sort_index(axis=1),
										displayable_metadata[self.slibs].sort_index(axis=1),
										on=displayable_metadata.index)

		return displayable_metadata


	def __write_to_excel(self, dataframe, sheet_name, auto_format = False):
		excel_workbook = load_workbook(self.dstpath)
		writer = pd.ExcelWriter(self.dstpath, engine = 'openpyxl')
		writer.book = excel_workbook
		dataframe.to_excel(writer, sheet_name = sheet_name)
		sheet = writer.sheets[sheet_name]
		sheet.column_dimensions['A'].width = 30
		writer.save()
		writer.close()
		if auto_format:
			self.__sheet_autoformat(sheet_name)

	def __sheet_autoformat(self, sheet_name):
		excel_workbook = load_workbook(self.dstpath)
		sheet = excel_workbook[sheet_name]

		for column in sheet.columns:
			length = max(len(myutils.as_text(cell.value)) for cell in column)
			sheet.column_dimensions[myutils.index_to_letter(column[0].column)].width = length
		excel_workbook.save(self.dstpath)

	def __delete_standard_sheet(self):
		excel_workbook = load_workbook(self.dstpath)
		if len(excel_workbook.sheetnames) > 1:
			del excel_workbook['Sheet']
			excel_workbook.save(self.dstpath)