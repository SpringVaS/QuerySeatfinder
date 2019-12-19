from __future__ import annotations
from abc import ABC, abstractmethod
from random import randrange
from typing import List

import requests
import json
import functools

from openpyxl import Workbook, load_workbook

import pandas as pd
import numpy as np

import myutils

URL = "https://seatfinder.bibliothek.kit.edu/karlsruhe/getdata.php"

class Subject(ABC):
	"""
	The Subject interface declares a set of methods for managing subscribers.
	"""

	@abstractmethod
	def attach(self, observer: Observer) -> None:
		"""
		Attach an observer to the subject.
		"""
		pass

	@abstractmethod
	def detach(self, observer: Observer) -> None:
		"""
		Detach an observer from the subject.
		"""
		pass

	@abstractmethod
	def notify(self) -> None:
		"""
		Notify all observers about an event.
		"""
		pass


class Observer(ABC):
	"""
	The Observer interface declares the update method, used by subjects.
	"""

	@abstractmethod
	def update(self, subject: Subject) -> None:
		"""
		Receive update from subject.
		"""
		pass


class Model(Subject):

	_state: int = None
	
	"""
	For the sake of simplicity, the Subject's state, essential to all
	subscribers, is stored in this variable.
	"""

	_observers: List[Observer] = []
	"""
	List of subscribers. In real life, the list of subscribers can be stored
	more comprehensively (categorized by event type, etc.).
	"""

	def attach(self, observer: Observer) -> None:
		print("Subject: Attached an observer.")
		self._observers.append(observer)

	def detach(self, observer: Observer) -> None:
		self._observers.remove(observer)

	"""
	The subscription management methods.
	"""

	def notify(self) -> None:
		"""
		Trigger an update in each subscriber.
		"""

		print("Subject: Notifying observers...")
		for observer in self._observers:
			observer.update(self)

	def __init__(self, url_seatfinder, dstpath):
		self.url_sf = url_seatfinder
		self.dstpath = dstpath
		# Reading halls LS
		self.mainlib = {'LSG',      # Gesellschaftswissenschaften
						'LSM',      # Medienzentrum
						'LST',      # Technik
						'LSN',      # Naturwissenschaften
						'LSW',      # Wiwi und Informatik
						'LBS'}         # Lehrbuchsammlung
						#'BIB-N'}   # KIT-Bibliothek Nord. Am Campus Nord
		# Specific Libraires
		self.slibs =   {'FBC',      # Chemie
						'FBP',      # Physik
						'LAF',      # Lernzentrum am Fasanenschloesschen}
						'FBA',      # Architektur
						'FBI',      # Informatik
						'FBM'}     # Mathematik
						#'FBW'}     # Wiwi - zur Zeit geschlossen

		self.vtypes = {'location[]', 'sublocs'}

		self.timeSeriesKeys = {'seatestimate' : 'occupied_seats', 'manualcount' : 'occupied_seats'}

		self.libMetadata = self.__getStaticLibData()
		self.query_progress = 0

		  # Create Exel File
		workbook = Workbook()
		workbook.save(dstpath)

		self.writeInfoToExcel(self.libMetadata, "Libraires", True)
		self.__deleteStandartSheet()

	def __del__(self):
		pass

	def getInfo(self, kind, timebegin, timeend):
		data = self.__queryServer(kind, [self.mainlib, self.slibs], timebegin, timeend)
		resampled = {}
		self.query_progress = 0
		self.notify()
		ctn = 0
		for location in data:
			locationData = location[kind]
			locationKey = next(iter(locationData))
			pddf =  self.__parseTimeSeries(kind, locationData, locationKey)

			oldestTime = pddf.iloc[0].name 
			while oldestTime > timebegin:
				reload_data = self.__queryServer(kind, locationKey, timebegin, oldestTime)
				for locationr in reload_data:
					pddfr =  self.__parseTimeSeries(kind, locationr[kind], next(iter(locationr[kind])))
					pddf = pddf.append(pddfr)
					oldestTime = pddfr.iloc[0].name
			pddf = pddf[pddf.index >= timebegin]

			sampleddf = pddf.resample('15Min').mean()
			resampled[locationKey] = sampleddf.round()
			ctn += 1
			self.query_progress = (ctn / len(data)) * 100
			self.notify()
		
		combinedData = functools.reduce(lambda left,right: 
			pd.merge(left,right,on='timestamp', how = 'outer').fillna(0), 
			resampled.values())
		combinedData = combinedData.sort_index(ascending = False)
		return combinedData

	def __getStaticLibData(self):
		queryparams =    {'location[]'   : [self.mainlib, self.slibs], 
						 'sublocs'      : 1,
						 'values'       : 'location',
						 'before'       : 'now'}
		r = requests.get(url = self.url_sf, params = queryparams)
		data = r.json()
		# write back the json from server to local hard drive for debugging purposes
		with open('staticlibdata.json', 'w+') as ld:
			ld.write(r.text)

		"""data = {}
		with open('staticlibdata.json', 'r') as datafile:
			data = json.load(datafile)"""


		callbacks = {'timestamp' : self.__parseTimeStamps, 'opening_hours' : self.__parseOpeningHours}

		staticlibdata = {}
		for location in data:
			metadata = location['location']
			locationKey = next(iter(metadata))
			parsedMetaInfo = metadata[locationKey][0]
			for key in parsedMetaInfo.keys():
				if key in callbacks.keys():
					parsedMetaInfo[key] = callbacks[key](parsedMetaInfo[key])
			staticlibdata[locationKey] = parsedMetaInfo

		return pd.DataFrame(staticlibdata)

	def getQueryProgess(self):
		return self.query_progress

	def writeInfoToExcel(self, dataframe, sheet_name, autoFormat = False):
		excel_workbook = load_workbook(self.dstpath)
		writer = pd.ExcelWriter(self.dstpath, engine = 'openpyxl')
		writer.book = excel_workbook
		dataframe.to_excel(writer, sheet_name = sheet_name)
		sheet = writer.sheets[sheet_name]
		sheet.column_dimensions['A'].width = 30
		writer.save()
		writer.close()
		if autoFormat:
			self.__autoFormatSheet(sheet_name)

	def __autoFormatSheet(self, sheet_name):
		excel_workbook = load_workbook(self.dstpath)
		sheet = excel_workbook[sheet_name]

		for column in sheet.columns:
			length = max(len(myutils.as_text(cell.value)) for cell in column)
			sheet.column_dimensions[myutils.letterFromIndex(column[0].column)].width = length
		excel_workbook.save(self.dstpath)

	def __deleteStandartSheet(self):
		excel_workbook = load_workbook(self.dstpath)
		if len(excel_workbook.sheetnames) > 1:
			del excel_workbook['Sheet']
			excel_workbook.save(self.dstpath)

	def __parseTimeSeries(self, kind, data, locationKey):
		data_list = data[locationKey]
		timestamp_list = [self.__parseTimeStamps(pointInTime['timestamp']) for pointInTime in data_list]
		value_list = [pointInTime[self.timeSeriesKeys[kind]] for pointInTime in data_list]

		value_dict = {'timestamp' : timestamp_list, locationKey : value_list}
		timeSeriesDataFrame = pd.DataFrame(value_dict)
		timeSeriesDataFrame = timeSeriesDataFrame.set_index(['timestamp'])
		return timeSeriesDataFrame

	def __parseOpeningHours(self, data):
		ohlist = []
		# weekly opening hours
		#print(data[keylist[1]])
		for interval in data['weekly_opening_hours']:
			openingHours_str = ""
			opening = self.__parseTimeStamps(interval[0])
			closing = self.__parseTimeStamps(interval[1])
			weekday_opening = opening.strftime("%A")
			weekday_closing = closing.strftime("%A")
			time_opening = opening.strftime("%H:%M")
			time_closing = closing.strftime("%H:%M")
			if (weekday_opening == weekday_closing):
				openingHours_str = weekday_opening + ": " + time_opening + " - " + time_closing
			else:
				openingHours_str = weekday_opening + ": " + time_opening + " - " + weekday_closing + ": " + time_closing
			#print(opening)
			ohlist.append(openingHours_str)
		return ohlist

	def __parseTimeStamps(self, data):
		#sheet.write(row, column, str(data))
		if (not isinstance(data, dict)):
			return data
		keylist = list(data.keys());
		if (keylist[0] == 'date'):
			return pd.Timestamp(data['date'])

	def __queryServer(self, kind, location_list, timebegin, timeend):
		queryparams =   {'location[]'   : location_list, 
						 'sublocs'      : 0,
						 'values'       : kind,
						 'before'       : timeend,
						 'limit'        : 1000}
		r = requests.get(url = self.url_sf, params = queryparams)
		data = r.json()
		# write back the json from server to local hard drive for debugging purposes
		with open('libdata.json', 'w+') as ld:
			ld.write(r.text)
		return data

	# TO DO: Get Weekday Names from Python environment
	def __defineWeekdayNames(self):
		mon = datetime.datetime.strptime("2013-05-13 09:00:00.000000", "%Y-%m-%d %H:%M:%S.%f")
		print((mon + 1).strftime("%A"))