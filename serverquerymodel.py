from __future__ import annotations
from abc import ABC, abstractmethod
from typing import List

import requests
import json
from functools import partial, reduce

from warnings import warn

from openpyxl import Workbook, load_workbook

import pandas as pd
import numpy as np

import myutils

URL = "https://seatfinder.bibliothek.kit.edu/karlsruhe/getdata.php"

class Subject(ABC):
	"""
	The Subject interface declares a set of methods for managing subscribers.
	"""

	@abstractmethod
	def attach(self, observer: Observer) -> None:
		"""
		Attach an observer to the subject.
		"""
		pass

	@abstractmethod
	def detach(self, observer: Observer) -> None:
		"""
		Detach an observer from the subject.
		"""
		pass

	@abstractmethod
	def notify(self) -> None:
		"""
		Notify all observers about an event.
		"""
		pass


class Observer(ABC):
	"""
	The Observer interface declares the update method, used by subjects.
	"""

	@abstractmethod
	def update(self, subject: Subject) -> None:
		"""
		Receive update from subject.
		"""
		pass


class Model(Subject):

	_state: int = None
	
	"""
	For the sake of simplicity, the Subject's state, essential to all
	subscribers, is stored in this variable.
	"""

	_observers: List[Observer] = []
	"""
	List of subscribers. In real life, the list of subscribers can be stored
	more comprehensively (categorized by event type, etc.).
	"""

	def attach(self, observer: Observer) -> None:
		print("Subject: Attached an observer.")
		self._observers.append(observer)

	def detach(self, observer: Observer) -> None:
		self._observers.remove(observer)

	"""
	The subscription management methods.
	"""

	def notify(self) -> None:
		"""
		Trigger an update in each subscriber.
		"""

		for observer in self._observers:
			observer.update(self)

	def __init__(self, url_seatfinder, dstpath):
		self.url_sf = url_seatfinder
		self.dstpath = dstpath
		# Reading halls LS
		self.mainlib = {'LSG',      # Gesellschaftswissenschaften
						'LSM',      # Medienzentrum
						'LST',      # Technik
						'LSN',      # Naturwissenschaften
						'LSW',      # Wiwi und Informatik
						'LBS'}      # Lehrbuchsammlung
						#'BIB-N'}   # KIT-Bibliothek Nord. Am Campus Nord
		# Specific Libraires
		self.slibs =   {'FBC',      # Chemie
						'FBP',      # Physik
						'LAF',      # Lernzentrum am Fasanenschloesschen}
						'FBA',      # Architektur
						'FBI',      # Informatik
						'FBM'}     # Mathematik
						#'FBW'}     # Wiwi - zur Zeit geschlossen

		self.allLocationsOnCampus = [*(self.mainlib), *(self.slibs)]


		self.resampling_interval = '15Min'

		self.vtypes = {'location[]', 'sublocs'}

		self.timeSeriesKeys = {'seatestimate' : 'occupied_seats', 'manualcount' : 'occupied_seats'}

		self.sampling_methods = ['Mean', 'Gauss']
		self.selected_sampling_method = 'mean'

		self.libMetadata = self.__get_static_lib_data()
		self.query_progress = 0

		  # Create Exel File
		workbook = Workbook()
		workbook.save(dstpath)

		self.write_to_excel(self.libMetadata, "Libraires", True)
		self.__delete_standard_sheet()

	def __del__(self):
		pass

	def get_sampling_method_options(self):
		return self.sampling_methods

	def set_sampling_method(self, value):
		self.selected_sampling_method = value


	def get_info(self, kind, timebegin, timeend):
		resampled = {}
		location_index = 0
		total_locations = len(self.allLocationsOnCampus)
		for location_id in self.allLocationsOnCampus:
			self.__update_progress((location_index / total_locations) * 100)
			oldestTime = timeend
			rawdata = pd.DataFrame()
			while oldestTime > timebegin:
				reload_data = self.__get_info_for_location_from_server(location_id, kind, timebegin, oldestTime)
				newOldestTime = oldestTime
				if (reload_data.size > 0):
					newOldestTime = reload_data.iloc[0].name
				else:
					break
				# assure strictly monotonous decline
				if (not (newOldestTime < oldestTime)):
					break
				
				# keep data sorted
				rawdata = rawdata.append(reload_data.sort_index(ascending = False))

				time_progress = newOldestTime if newOldestTime > timebegin else timebegin
				self.__update_progress(self.get_progress() + 
					((oldestTime - time_progress) / (timeend - timebegin)) * (100 / (total_locations)))
				oldestTime = newOldestTime

			#rawdata = rawdata[rawdata.index >= timebegin]
			rawdata = rawdata.truncate(after = timebegin)

			location_data = self.__resample(rawdata)
			
			#location_data = rawdata.resample(self.resampling_interval).mean()

			resampled[location_id] = location_data.round()
			location_index += 1

			# dataframe output for aggregation description
			#self.write_to_excel(rawdata, 'raw ' + str(location_id))
			#self.write_to_excel(location_data, 'resampled ' + str(location_id))


		self.__update_progress(100)

		combinedData = reduce(lambda left,right: 
			pd.merge(left,right,on='timestamp', how = 'outer').fillna(0), 
			resampled.values())
		combinedData = combinedData.sort_index(ascending = False)
		return combinedData

	def __resample(self, data):
		""" Without additional parameters, the pandas datframe resample function
			aggregates the values between the two labels. To simulate a moving average
			the actual label ticks are moved half the resampling interval.
			The label names are also offset half the resampling interval.
		"""
		timedelta = pd.Timedelta(self.resampling_interval)
		interval_seconds = str(timedelta.seconds) + 'S'
		offset_delta = pd.Timedelta(0)
		if (timedelta < pd.Timedelta('1D')):
			offset_delta = timedelta / 2

		"""
		Apply custom aggregarion function. Use functools to include parameter.
		"""
		if (timedelta >= pd.Timedelta('1W')):
			''' Use monday as base label
			'''
			resampler = data.resample(self.resampling_interval, label = 'left', loffset='1D')

		elif (offset_delta.seconds == 0):
			resampler = data.resample(self.resampling_interval)
		else:
			resampler = data.resample(interval_seconds, base=offset_delta.seconds, loffset=offset_delta)

		
		#resampler = data.resample(self.resampling_interval)

		location_data = pd.DataFrame()

		if (timedelta >= pd.Timedelta('1D') or self.selected_sampling_method == 'Mean'):
			location_data = resampler.mean()
		elif (self.selected_sampling_method == 'Gauss'):
			sigma = myutils.calculate_derivation(offset_delta.seconds, 0.2)
			location_data = myutils.resample_gaussian(resampler, sigma)

		location_data = location_data.sort_index(ascending = False)
		return location_data



	def set_resampling_interval(self, value):
		self.resampling_interval = value

	def __get_info_for_location_from_server(self, location_id, kind, timebegin, timeend):
		data = self.__query_server(kind, location_id, timebegin, timeend)
		pddf = pd.DataFrame()
		if (len(data) > 0):
			assert (len(data) == 1), ("Please review location_id parameter " + str(len(data)))
			for location in data:
				locationData = location[kind]
				pddf = pddf.append(self.__parse_timeseries(kind, locationData, location_id))
		return pddf


	def __get_static_lib_data(self):
		queryparams =    {'location[]'   : [self.mainlib, self.slibs], 
						 'sublocs'      : 1,
						 'values'       : 'location',
						 'before'       : 'now'}
		data = {}
		try:
			r = requests.get(url = self.url_sf, params = queryparams)
			data = r.json()
			# write back the json from server to local hard drive for debugging purposes
			with open('staticlibdata.json', 'w+') as ld:
				ld.write(r.text)
			print('Load data from server and write back to file')
		except requests.ConnectionError as e:
			print('Load data from file')
			with open('staticlibdata.json', 'r') as datafile:
				data = json.load(datafile)


		callbacks = {	'timestamp' 	: self.__parse_timestamps, 
						'opening_hours' : self.__parse_openinghours}

		staticlibdata = {}
		for location in data:
			metadata = location['location']
			locationKey = next(iter(metadata))
			parsedMetaInfo = metadata[locationKey][0]
			for key in parsedMetaInfo.keys():
				if key in callbacks.keys():
					parsedMetaInfo[key] = callbacks[key](parsedMetaInfo[key])
			staticlibdata[locationKey] = parsedMetaInfo

		return pd.DataFrame(staticlibdata)

	def get_progress(self):
		return self.query_progress

	def write_to_excel(self, dataframe, sheet_name, auto_format = False):
		excel_workbook = load_workbook(self.dstpath)
		writer = pd.ExcelWriter(self.dstpath, engine = 'openpyxl')
		writer.book = excel_workbook
		dataframe.to_excel(writer, sheet_name = sheet_name)
		sheet = writer.sheets[sheet_name]
		sheet.column_dimensions['A'].width = 30
		writer.save()
		writer.close()
		if auto_format:
			self.__sheet_autoformat(sheet_name)

	def __sheet_autoformat(self, sheet_name):
		excel_workbook = load_workbook(self.dstpath)
		sheet = excel_workbook[sheet_name]

		for column in sheet.columns:
			length = max(len(myutils.as_text(cell.value)) for cell in column)
			sheet.column_dimensions[myutils.index_to_letter(column[0].column)].width = length
		excel_workbook.save(self.dstpath)

	def __delete_standard_sheet(self):
		excel_workbook = load_workbook(self.dstpath)
		if len(excel_workbook.sheetnames) > 1:
			del excel_workbook['Sheet']
			excel_workbook.save(self.dstpath)

	def __parse_timeseries(self, kind, data, locationKey):
		data_list = data[locationKey]
		timestamp_list = [self.__parse_timestamps(pointInTime['timestamp']) for pointInTime in data_list]
		value_list = [pointInTime[self.timeSeriesKeys[kind]] for pointInTime in data_list]

		value_dict = {'timestamp' : timestamp_list, locationKey : value_list}
		timeSeriesDataFrame = pd.DataFrame(value_dict)
		timeSeriesDataFrame = timeSeriesDataFrame.set_index(['timestamp'])
		return timeSeriesDataFrame

	def __parse_openinghours(self, data):
		ohlist = []
		# weekly opening hours
		#print(data[keylist[1]])
		for interval in data['weekly_opening_hours']:
			openingHours_str = ""
			opening = self.__parse_timestamps(interval[0])
			closing = self.__parse_timestamps(interval[1])
			weekday_opening = opening.strftime("%A")
			weekday_closing = closing.strftime("%A")
			time_opening = opening.strftime("%H:%M")
			time_closing = closing.strftime("%H:%M")
			if (weekday_opening == weekday_closing):
				openingHours_str = weekday_opening + ": " + time_opening + " - " + time_closing
			else:
				openingHours_str = weekday_opening + ": " + time_opening + " - " + weekday_closing + ": " + time_closing
			#print(opening)
			ohlist.append(openingHours_str)
		return ohlist

	def __parse_timestamps(self, data):
		#sheet.write(row, column, str(data))
		if (not isinstance(data, dict)):
			return data
		keylist = list(data.keys());
		if (keylist[0] == 'date'):
			return pd.Timestamp(data['date'])

	def __query_server(self, kind, location_list, timebegin, timeend):
		queryparams =   {'location[]'   : location_list, 
						 'sublocs'      : 0,
						 'values'       : kind,
						 'before'       : timeend,
						 'limit'        : 1000}
		data = {}
		try:
			r = requests.get(url = self.url_sf, params = queryparams)
			data = r.json()
			# write back the json from server to local hard drive for debugging purposes
			with open('libdata.json', 'w+') as ld:
				ld.write(r.text)
		except requests.ConnectionError as e:
			warn('No server connection')

		return data

	def __update_progress(self, value):
		self.query_progress = value
		self.notify()

	# TO DO: Get Weekday Names from Python environment
	def __defineWeekdayNames(self):
		mon = datetime.datetime.strptime("2013-05-13 09:00:00.000000", "%Y-%m-%d %H:%M:%S.%f")
		print((mon + 1).strftime("%A"))