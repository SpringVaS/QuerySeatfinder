from __future__ import annotations
from abc import ABC, abstractmethod
from typing import List

from functools import reduce
from openpyxl import Workbook, load_workbook

import pandas as pd

import myutils
from servercomm import ServerCommunication
from dataprocessing import DataProcessor

class Subject(ABC):
	"""
	The Subject interface declares a set of methods for managing subscribers.
	"""

	@abstractmethod
	def attach(self, observer: Observer) -> None:
		"""
		Attach an observer to the subject.
		"""
		pass

	@abstractmethod
	def detach(self, observer: Observer) -> None:
		"""
		Detach an observer from the subject.
		"""
		pass

	@abstractmethod
	def notify(self) -> None:
		"""
		Notify all observers about an event.
		"""
		pass


class Observer(ABC):
	"""
	The Observer interface declares the update method, used by subjects.
	"""

	@abstractmethod
	def update(self, subject: Subject) -> None:
		"""
		Receive update from subject.
		"""
		pass


class Model(Subject):

	_state: int = None
	
	"""
	For the sake of simplicity, the Subject's state, essential to all
	subscribers, is stored in this variable.
	"""

	_observers: List[Observer] = []
	"""
	List of subscribers. In real life, the list of subscribers can be stored
	more comprehensively (categorized by event type, etc.).
	"""

	def attach(self, observer: Observer) -> None:
		print("Subject: Attached an observer.")
		self._observers.append(observer)

	def detach(self, observer: Observer) -> None:
		self._observers.remove(observer)

	"""
	The subscription management methods.
	"""

	def notify(self) -> None:
		"""
		Trigger an update in each subscriber.
		"""

		for observer in self._observers:
			observer.update(self)

	def __init__(self, url_seatfinder, dstpath):
		self.dstpath = dstpath

		self.server = ServerCommunication(url_seatfinder)

		# Reading halls LS
		self.mainlib = {'LSG',      # Gesellschaftswissenschaften
						'LSM',      # Medienzentrum
						'LST',      # Technik
						'LSN',      # Naturwissenschaften
						'LSW',      # Wiwi und Informatik
						'LBS'}      # Lehrbuchsammlung
						#'BIB-N'}   # KIT-Bibliothek Nord. Am Campus Nord
		# Specific Libraires
		self.slibs =   {'FBC',      # Chemie
						'FBP',      # Physik
						'LAF',      # Lernzentrum am Fasanenschloesschen}
						'FBA',      # Architektur
						'FBI',      # Informatik
						'FBM'}     # Mathematik
						#'FBW'}     # Wiwi - zur Zeit geschlossen

		self.all_libs = [*(self.mainlib), *(self.slibs)]


		self.resampling_interval = '15Min'

		self.sampling_methods = ['Mean', 'Gauss']
		self.selected_sampling_method = self.sampling_methods[0]

		self.libs_metadata = self.server.get_static_lib_data(self.all_libs)

		self.data_processor = DataProcessor(self.mainlib, self.slibs, self.libs_metadata)

		self.query_progress = 0

		  # Create Exel File
		workbook = Workbook()
		workbook.save(dstpath)

		self.__write_to_excel(self.__displayable_lib_metadata(), "Libraries", True)
		self.__delete_standard_sheet()


	def set_resampling_interval(self, value):
		self.resampling_interval = value


	def get_sampling_method_options(self):
		return self.sampling_methods

	def set_sampling_method(self, value):
		self.selected_sampling_method = value

	def get_progress(self):
		return self.query_progress

	def get_info(self, kind, timebegin, timeend):
		resampled = {}
		location_index = 0
		total_locations = len(self.all_libs)
		for location_id in self.all_libs:
			self.__update_progress((location_index / total_locations) * 95)
			oldestTime = timeend
			rawdata = pd.DataFrame()
			while oldestTime > timebegin:
				reload_data = self.server.get_info_for_location_from_server(location_id, kind, timebegin, oldestTime)
				newOldestTime = oldestTime
				if (reload_data.size > 0):
					newOldestTime = reload_data.iloc[0].name
				else:
					break
				# assure strictly monotonous decline
				if (not (newOldestTime < oldestTime)):
					break
				
				# keep data sorted
				rawdata = rawdata.append(reload_data.sort_index(ascending = False))

				time_progress = newOldestTime if newOldestTime > timebegin else timebegin
				self.__update_progress(self.get_progress() + 
					((oldestTime - time_progress) / (timeend - timebegin)) * (95 / (total_locations)))
				oldestTime = newOldestTime

			#rawdata = rawdata[rawdata.index >= timebegin]
			rawdata = rawdata.truncate(after = timebegin)

			location_data = self.data_processor.resample(rawdata, self.resampling_interval, 
				self.selected_sampling_method)
			
			#location_data = rawdata.resample(self.resampling_interval).mean()

			resampled[location_id] = location_data.round()
			location_index += 1

		self.__update_progress(95)

		combinedData = reduce(lambda left,right: 
			pd.merge(left,right,on='timestamp', how = 'outer').fillna(0), 
			resampled.values())
		combinedData.sort_index(ascending = False, inplace = True)
		return combinedData

	def seat_estimate_and_pressure_to_excel(self, timebegin, timeend):
		occupancy = self.get_info('seatestimate', timebegin, timeend)
		mainlib_pressure = self.data_processor.compute_pressure(occupancy)
		grouped_occupancy = self.__grouped_seat_info(occupancy)
		self.__write_to_excel(grouped_occupancy, "Occupancy")
		self.__write_to_excel(mainlib_pressure, "Proportional occupancy")
		self.__update_progress(100)

	def __displayable_lib_metadata(self):
		displayable_metadata = self.libs_metadata.copy(deep = True)
		for location_id in self.libs_metadata.keys():
			entry = displayable_metadata[location_id]['opening_hours']
			ohlist_display = []
			for interval in entry:
				openingHours_str = ""
				weekday_opening = interval.left.strftime("%A")
				weekday_closing = interval.right.strftime("%A")
				time_opening = interval.left.strftime("%H:%M")
				time_closing = interval.right.strftime("%H:%M")
				if (weekday_opening == weekday_closing):
					openingHours_str = weekday_opening + ": " + time_opening + " - " + time_closing
				else:
					openingHours_str = weekday_opening + ": " + time_opening + " - " + weekday_closing + ": " + time_closing
				ohlist_display.append(openingHours_str)
			displayable_metadata[location_id]['opening_hours'] = ohlist_display

		displayable_metadata = pd.merge(displayable_metadata[self.mainlib].sort_index(axis=1),
										displayable_metadata[self.slibs].sort_index(axis=1),
										on=displayable_metadata.index)

		return displayable_metadata

	def __grouped_seat_info(self, data):
		mainlib = data[self.mainlib].sum(axis=1)
		mainlib.name = "KIT-BIB"
		merged = pd.merge(mainlib, data[self.slibs].sort_index(axis=1), on='timestamp')
		return merged


	def __write_to_excel(self, dataframe, sheet_name, auto_format = False):
		excel_workbook = load_workbook(self.dstpath)
		writer = pd.ExcelWriter(self.dstpath, engine = 'openpyxl')
		writer.book = excel_workbook
		dataframe.to_excel(writer, sheet_name = sheet_name)
		sheet = writer.sheets[sheet_name]
		sheet.column_dimensions['A'].width = 30
		writer.save()
		writer.close()
		if auto_format:
			self.__sheet_autoformat(sheet_name)

	def __sheet_autoformat(self, sheet_name):
		excel_workbook = load_workbook(self.dstpath)
		sheet = excel_workbook[sheet_name]

		for column in sheet.columns:
			length = max(len(myutils.as_text(cell.value)) for cell in column)
			sheet.column_dimensions[myutils.index_to_letter(column[0].column)].width = length
		excel_workbook.save(self.dstpath)

	def __delete_standard_sheet(self):
		excel_workbook = load_workbook(self.dstpath)
		if len(excel_workbook.sheetnames) > 1:
			del excel_workbook['Sheet']
			excel_workbook.save(self.dstpath)

	def __update_progress(self, value):
		self.query_progress = value
		self.notify()